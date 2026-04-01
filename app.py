from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtCore import QSignalBlocker, Qt, QUrl, Signal
from PySide6.QtGui import QColor, QKeySequence, QPainter, QPainterPath, QPen, QShortcut
from PySide6.QtMultimedia import QAudioBuffer, QAudioDecoder, QAudioFormat, QAudioOutput, QMediaPlayer
from PySide6.QtWidgets import (
    QApplication,
    QButtonGroup,
    QCheckBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QGraphicsDropShadowEffect,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QLineEdit,
    QPlainTextEdit,
    QRadioButton,
    QScrollArea,
    QSizePolicy,
    QVBoxLayout,
    QWidget,
)

from question_config import QUESTION_SECTIONS, REMARK_KEY


SUPPORTED_AUDIO_EXTENSIONS = {
    ".aac",
    ".amr",
    ".flac",
    ".m4a",
    ".mp3",
    ".ogg",
    ".opus",
    ".wav",
    ".wma",
}

THEMES = {
    "light": {
        "window_start": "#f4f7fb",
        "window_end": "#e8eef8",
        "panel_bg": "rgba(255, 255, 255, 0.72)",
        "panel_alt_bg": "rgba(255, 255, 255, 0.58)",
        "card_bg": "rgba(255, 255, 255, 0.86)",
        "card_border": "rgba(124, 146, 181, 0.22)",
        "text_primary": "#162033",
        "text_muted": "#5f6f8f",
        "accent": "#1f6feb",
        "accent_hover": "#175fd1",
        "secondary_bg": "rgba(223, 231, 245, 0.72)",
        "secondary_hover": "rgba(214, 225, 242, 0.96)",
        "input_bg": "rgba(250, 252, 255, 0.92)",
        "input_border": "rgba(154, 172, 204, 0.35)",
        "shadow": "#adc1e6",
        "shadow_alpha": 80,
        "slider_groove": "rgba(173, 190, 220, 0.45)",
    },
    "dark": {
        "window_start": "#0f1726",
        "window_end": "#172235",
        "panel_bg": "rgba(18, 28, 44, 0.76)",
        "panel_alt_bg": "rgba(18, 28, 44, 0.66)",
        "card_bg": "rgba(24, 36, 56, 0.88)",
        "card_border": "rgba(138, 166, 214, 0.22)",
        "text_primary": "#f4f7ff",
        "text_muted": "#a9b6cf",
        "accent": "#6ea8ff",
        "accent_hover": "#5c97f4",
        "secondary_bg": "rgba(38, 54, 80, 0.9)",
        "secondary_hover": "rgba(50, 68, 98, 0.96)",
        "input_bg": "rgba(18, 28, 44, 0.95)",
        "input_border": "rgba(112, 139, 184, 0.35)",
        "shadow": "#08111f",
        "shadow_alpha": 160,
        "slider_groove": "rgba(90, 114, 156, 0.54)",
    },
}


def format_ms(ms: int) -> str:
    seconds = max(ms, 0) // 1000
    minute, second = divmod(seconds, 60)
    hour, minute = divmod(minute, 60)
    if hour:
        return f"{hour:02d}:{minute:02d}:{second:02d}"
    return f"{minute:02d}:{second:02d}"


def normalize_audio_key(name: str) -> str:
    normalized = Path(str(name).strip()).stem.strip().lower()
    match = re.search(r"([a-z0-9]+-\d+)", normalized)
    if match:
        return match.group(1)
    return normalized


class WaveformSeekBar(QWidget):
    seekRequested = Signal(float)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setMinimumHeight(40)
        self.setCursor(Qt.PointingHandCursor)
        self.waveform_points: list[tuple[float, float]] = []
        self.progress_ratio = 0.0
        self.dragging = False
        self.colors = {
            "background": QColor("#dbe5f4"),
            "played": QColor("#1f6feb"),
            "upcoming": QColor("#8ba5cd"),
            "line": QColor("#ffffff"),
        }

    def set_waveform(self, points: list[tuple[float, float]]) -> None:
        self.waveform_points = points or [(-0.18, 0.18)] * 180
        self.update()

    def set_progress(self, position_ms: int, duration_ms: int) -> None:
        if duration_ms <= 0:
            self.progress_ratio = 0.0
        else:
            self.progress_ratio = max(0.0, min(1.0, position_ms / duration_ms))
        if not self.dragging:
            self.update()

    def set_theme(self, *, background: str, played: str, upcoming: str, line: str) -> None:
        self.colors = {
            "background": QColor(background),
            "played": QColor(played),
            "upcoming": QColor(upcoming),
            "line": QColor(line),
        }
        self.update()

    def mousePressEvent(self, event) -> None:  # type: ignore[override]
        if event.button() == Qt.LeftButton:
            self.dragging = True
            self._emit_seek(event.position().x())

    def mouseMoveEvent(self, event) -> None:  # type: ignore[override]
        if self.dragging:
            self._emit_seek(event.position().x())

    def mouseReleaseEvent(self, event) -> None:  # type: ignore[override]
        if event.button() == Qt.LeftButton:
            self.dragging = False
            self._emit_seek(event.position().x())

    def paintEvent(self, event) -> None:  # type: ignore[override]
        del event
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        rect = self.rect().adjusted(0, 4, 0, -4)
        path = QPainterPath()
        path.addRoundedRect(rect, 12, 12)
        painter.fillPath(path, self.colors["background"])

        points = self.waveform_points or [(-0.18, 0.18)] * 180
        count = len(points)
        if count == 0:
            return

        bar_width = max(2.0, rect.width() / max(count * 1.9, 1))
        step = rect.width() / count
        center_y = rect.center().y()
        progress_x = rect.left() + rect.width() * self.progress_ratio

        played_pen = QPen(self.colors["played"], bar_width, Qt.SolidLine, Qt.RoundCap)
        upcoming_pen = QPen(self.colors["upcoming"], bar_width, Qt.SolidLine, Qt.RoundCap)

        for index, (min_value, max_value) in enumerate(points):
            x = rect.left() + (index + 0.5) * step
            clamped_min = max(-1.0, min(1.0, min_value))
            clamped_max = max(-1.0, min(1.0, max_value))
            y_top = center_y - clamped_max * (rect.height() / 2 - 4)
            y_bottom = center_y - clamped_min * (rect.height() / 2 - 4)
            if abs(y_bottom - y_top) < 4:
                y_top = center_y - 2
                y_bottom = center_y + 2
            painter.setPen(played_pen if x <= progress_x else upcoming_pen)
            painter.drawLine(x, y_top, x, y_bottom)

        painter.setPen(QPen(self.colors["line"], 2))
        painter.drawLine(progress_x, rect.top() + 2, progress_x, rect.bottom() - 2)

    def _emit_seek(self, x: float) -> None:
        width = max(1, self.width())
        ratio = max(0.0, min(1.0, x / width))
        self.progress_ratio = ratio
        self.update()
        self.seekRequested.emit(ratio)


class AnnotationWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("语音标注软件")
        self.resize(1520, 900)

        self.audio_files: list[Path] = []
        self.audio_index = -1
        self.audio_dir: Path | None = None
        self.json_dir = Path.cwd() / "json"
        self.json_dir.mkdir(exist_ok=True)

        self.player = QMediaPlayer(self)
        self.audio_output = QAudioOutput(self)
        self.player.setAudioOutput(self.audio_output)
        self.audio_output.setVolume(1.0)

        self.single_groups: dict[str, QButtonGroup] = {}
        self.multi_groups: dict[str, list[QCheckBox]] = {}
        self.question_labels: dict[str, str] = {}
        self.question_configs: dict[str, dict] = {}
        self.required_question_keys: set[str] = set()
        self.multi_exclusive_values: dict[str, str] = {}
        self.transcript_map: dict[str, dict[str, str]] = {}
        self.annotated_audio_ids: set[str] = set()
        self.preview_cache = "{}"
        self.annotator_name = ""
        self.current_theme = "light"
        self.shadow_targets: list[QWidget] = []
        self.waveform_decoder: QAudioDecoder | None = None
        self.waveform_decode_path: Path | None = None
        self.waveform_samples: list[float] = []
        self.shortcuts: list[QShortcut] = []

        self._build_ui()
        self._register_shortcuts()
        self._connect_player_signals()
        self.apply_theme()
        self.update_summary()

    def _build_ui(self) -> None:
        self.root = QWidget()
        self.root.setObjectName("appRoot")
        self.setCentralWidget(self.root)

        root_layout = QVBoxLayout(self.root)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setObjectName("mainScrollArea")
        root_layout.addWidget(scroll_area)

        content = QWidget()
        content.setObjectName("contentWidget")
        scroll_area.setWidget(content)

        main_layout = QVBoxLayout(content)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(8)

        hero_card, hero_layout = self.create_glass_card("heroCard", 10, 8, 10, 8)
        hero_layout.setSpacing(6)
        hero_row = QHBoxLayout()
        hero_row.setSpacing(12)
        hero_row.setAlignment(Qt.AlignTop)

        hero_text = QVBoxLayout()
        hero_text.setSpacing(2)

        app_title = QLabel("语音标注软件")
        app_title.setObjectName("appTitle")
        #app_subtitle = QLabel("清晰、轻快、适合连续标注的桌面工作台。")
        #app_subtitle.setObjectName("appSubtitle")
        self.audio_name_label = QLabel("当前音频：未加载")
        self.audio_name_label.setObjectName("audioNameLabel")
        self.annotation_status_label = QLabel("未标注")
        self.annotation_status_label.setObjectName("annotationStatusLabel")

        audio_header_row = QHBoxLayout()
        audio_header_row.setSpacing(8)
        audio_header_row.addWidget(self.audio_name_label)
        audio_header_row.addWidget(self.annotation_status_label)
        audio_header_row.addStretch()

        hero_text.addWidget(app_title)
        #hero_text.addWidget(app_subtitle)
        hero_text.addLayout(audio_header_row)

        inline_controls = QHBoxLayout()
        inline_controls.setSpacing(6)

        self.prev_button = QPushButton("上一条")
        self.prev_button.setProperty("variant", "secondary")
        self.prev_button.clicked.connect(self.play_previous)

        self.play_button = QPushButton("继续播放")
        self.play_button.setProperty("variant", "primary")
        self.play_button.clicked.connect(self.resume_playback)

        self.pause_button = QPushButton("暂停")
        self.pause_button.setProperty("variant", "secondary")
        self.pause_button.clicked.connect(self.pause_playback)

        self.next_button = QPushButton("下一条")
        self.next_button.setProperty("variant", "secondary")
        self.next_button.clicked.connect(self.play_next)

        for button in (self.prev_button, self.play_button, self.pause_button, self.next_button):
            button.setMinimumHeight(30)
            button.setMinimumWidth(82)
            inline_controls.addWidget(button)

        hero_actions = QHBoxLayout()
        hero_actions.setSpacing(8)
        hero_actions.addStretch()

        self.theme_button = QPushButton("切换夜间")
        self.theme_button.setCheckable(True)
        self.theme_button.setProperty("variant", "secondary")
        self.theme_button.setMinimumHeight(34)
        self.theme_button.setMinimumWidth(108)
        self.theme_button.clicked.connect(self.toggle_theme)

        self.open_folder_button = QPushButton("打开音频文件夹")
        self.open_folder_button.setProperty("variant", "secondary")
        self.open_folder_button.setMinimumHeight(34)
        self.open_folder_button.setMinimumWidth(122)
        self.open_folder_button.clicked.connect(self.choose_audio_folder)

        self.open_excel_button = QPushButton("导入 Excel")
        self.open_excel_button.setProperty("variant", "secondary")
        self.open_excel_button.setMinimumHeight(34)
        self.open_excel_button.setMinimumWidth(104)
        self.open_excel_button.clicked.connect(self.choose_excel_file)

        hero_actions.addLayout(inline_controls)
        hero_actions.addWidget(self.theme_button)
        hero_actions.addWidget(self.open_excel_button)
        hero_actions.addWidget(self.open_folder_button)

        hero_row.addLayout(hero_text, 1)
        hero_row.addLayout(hero_actions)
        hero_layout.addLayout(hero_row)

        transcript_row = QHBoxLayout()
        transcript_row.setSpacing(8)
        transcript_row.setAlignment(Qt.AlignTop)

        source_box, source_layout = self.create_glass_card("infoCard", 8, 6, 8, 6)
        source_box.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
        source_layout.setSpacing(4)
        source_row = QHBoxLayout()
        source_row.setSpacing(6)
        source_title = QLabel("原文:")
        source_title.setObjectName("miniSectionTitle")
        self.source_text = QPlainTextEdit()
        self.source_text.setObjectName("transcriptViewer")
        self.source_text.setReadOnly(True)
        self.source_text.setFixedHeight(46)
        self.source_text.setPlaceholderText("导入 Excel 后显示对应录音的原文。")
        source_row.addWidget(source_title)
        source_row.addWidget(self.source_text, 1)
        source_layout.addLayout(source_row)

        translation_box, translation_layout = self.create_glass_card("infoCard", 8, 6, 8, 6)
        translation_box.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
        translation_layout.setSpacing(4)
        translation_row = QHBoxLayout()
        translation_row.setSpacing(6)
        translation_title = QLabel("中文翻译:")
        translation_title.setObjectName("miniSectionTitle")
        self.translation_text = QPlainTextEdit()
        self.translation_text.setObjectName("transcriptViewer")
        self.translation_text.setReadOnly(True)
        self.translation_text.setFixedHeight(46)
        self.translation_text.setPlaceholderText("导入 Excel 后显示对应录音的中文翻译。")
        translation_row.addWidget(translation_title)
        translation_row.addWidget(self.translation_text, 1)
        translation_layout.addLayout(translation_row)

        transcript_row.addWidget(source_box, 1)
        transcript_row.addWidget(translation_box, 1)
        hero_layout.addLayout(transcript_row)
        main_layout.addWidget(hero_card)

        progress_card, progress_layout = self.create_glass_card("panelCard", 8, 6, 8, 6)
        progress_row = QHBoxLayout()
        progress_row.setSpacing(8)

        self.time_label = QLabel("00:00 / 00:00")
        self.time_label.setObjectName("timeLabel")
        self.time_label.setAlignment(Qt.AlignCenter)
        self.time_label.setMinimumWidth(110)

        self.waveform_bar = WaveformSeekBar()
        self.waveform_bar.seekRequested.connect(self._on_waveform_seek)

        progress_row.addWidget(self.time_label)
        progress_row.addWidget(self.waveform_bar, 1)
        progress_layout.addLayout(progress_row)
        main_layout.addWidget(progress_card)

        questions_shell, questions_shell_layout = self.create_glass_card("panelCard", 10, 8, 10, 8)
        questions_title = QLabel("标注维度")
        questions_title.setObjectName("sectionTitle")
        questions_shell_layout.addWidget(questions_title)

        questions_panel = QWidget()
        questions_panel.setObjectName("questionsPanel")
        self.questions_layout = QGridLayout(questions_panel)
        self.questions_layout.setContentsMargins(0, 0, 0, 0)
        self.questions_layout.setHorizontalSpacing(6)
        self.questions_layout.setVerticalSpacing(6)
        self._build_question_groups()

        questions_shell_layout.addWidget(questions_panel)
        main_layout.addWidget(questions_shell)

        bottom_row = QHBoxLayout()
        bottom_row.setSpacing(8)
        bottom_row.setAlignment(Qt.AlignTop)

        remark_card, remark_layout = self.create_glass_card("panelCard", 8, 5, 8, 5)
        remark_card.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
        remark_layout.setSpacing(4)
        remark_title = QLabel("描述音频")
        remark_title.setObjectName("miniSectionTitle")
        self.remark_edit = QPlainTextEdit()
        self.remark_edit.setObjectName("remarkEdit")
        self.remark_edit.setPlaceholderText("简要描述这段音频的内容、场景或特殊情况。")
        self.remark_edit.textChanged.connect(self.update_summary)
        self.remark_edit.setFixedHeight(38)
        remark_layout.addWidget(remark_title)
        remark_layout.addWidget(self.remark_edit)
        bottom_row.addWidget(remark_card, 7)

        action_card, action_layout = self.create_glass_card("panelCard", 8, 6, 8, 6)
        action_card.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
        action_grid = QGridLayout()
        action_grid.setContentsMargins(0, 0, 0, 0)
        action_grid.setHorizontalSpacing(6)
        action_grid.setVerticalSpacing(6)

        annotator_title = QLabel("标注人")
        annotator_title.setObjectName("compactFieldTitle")
        self.annotator_edit = QLineEdit()
        self.annotator_edit.setObjectName("annotatorEdit")
        self.annotator_edit.setPlaceholderText("填写标注人姓名")
        self.annotator_edit.textChanged.connect(self._update_annotator_name)
        self.annotator_edit.setFixedHeight(30)

        annotator_field = QWidget()
        annotator_layout = QHBoxLayout(annotator_field)
        annotator_layout.setContentsMargins(0, 0, 0, 0)
        annotator_layout.setSpacing(6)
        annotator_layout.addWidget(annotator_title)
        annotator_layout.addWidget(self.annotator_edit)

        self.preview_button = QPushButton("预览 JSON")
        self.preview_button.setProperty("variant", "secondary")
        self.preview_button.setMinimumHeight(30)
        self.preview_button.setMinimumWidth(118)
        self.preview_button.clicked.connect(self.show_preview_dialog)

        self.reset_button = QPushButton("重做")
        self.reset_button.setProperty("variant", "secondary")
        self.reset_button.setMinimumHeight(30)
        self.reset_button.setMinimumWidth(118)
        self.reset_button.clicked.connect(self.reset_current_annotation)

        self.submit_button = QPushButton("提交")
        self.submit_button.setProperty("variant", "primary")
        self.submit_button.setMinimumHeight(30)
        self.submit_button.setMinimumWidth(118)
        self.submit_button.clicked.connect(self.submit_annotation)

        action_grid.addWidget(self.preview_button, 0, 0)
        action_grid.addWidget(annotator_field, 0, 1)
        action_grid.addWidget(self.reset_button, 1, 0)
        action_grid.addWidget(self.submit_button, 1, 1)
        action_layout.addLayout(action_grid)
        bottom_row.addWidget(action_card, 3)

        main_layout.addLayout(bottom_row)
        self.apply_default_selections()

    def _build_question_groups(self) -> None:
        questions = [question for section in QUESTION_SECTIONS for question in section["questions"]]

        for index, question in enumerate(questions):
            self.question_configs[question["key"]] = question
            self.question_labels[question["key"]] = question["label"]
            if question.get("required"):
                self.required_question_keys.add(question["key"])
            if question["type"] == "multi" and question.get("exclusive_value"):
                self.multi_exclusive_values[question["key"]] = question["exclusive_value"]

            card, card_layout = self.create_glass_card("questionCard", 8, 6, 8, 6)
            card.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
            card_layout.setSpacing(3)

            label = QLabel(f"{question['label']}  ({question['key']})")
            label.setObjectName("questionTitle")
            label.setWordWrap(True)
            card_layout.addWidget(label)

            if question["type"] == "single":
                group = QButtonGroup(self)
                group.setExclusive(True)
                self.single_groups[question["key"]] = group
                for option in question["options"]:
                    button = QRadioButton(str(option["value"]))
                    button.setProperty("value", option["value"])
                    button.setToolTip(str(option["label"]))
                    button.setObjectName("optionButton")
                    button.toggled.connect(self.update_summary)
                    group.addButton(button)
                    card_layout.addWidget(button)
            else:
                checkboxes: list[QCheckBox] = []
                self.multi_groups[question["key"]] = checkboxes
                for option in question["options"]:
                    checkbox = QCheckBox(str(option["value"]))
                    checkbox.setProperty("value", option["value"])
                    checkbox.setToolTip(str(option["label"]))
                    checkbox.setObjectName("optionButton")
                    checkbox.stateChanged.connect(
                        lambda _state, key=question["key"], box=checkbox: self._handle_multi_change(key, box)
                    )
                    checkboxes.append(checkbox)
                    card_layout.addWidget(checkbox)

            row = index // 4
            column = index % 4
            self.questions_layout.addWidget(card, row, column)

        for column in range(4):
            self.questions_layout.setColumnStretch(column, 1)

    def _connect_player_signals(self) -> None:
        self.player.positionChanged.connect(self._sync_position)
        self.player.durationChanged.connect(self._sync_duration)
        self.player.mediaStatusChanged.connect(self._on_media_status_changed)

    def _register_shortcuts(self) -> None:
        shortcut_bindings = [
            ("Ctrl+Left", self.play_previous),
            ("Ctrl+Right", self.play_next),
            ("Space", self.toggle_playback),
            ("Return", self.submit_annotation),
            ("Enter", self.submit_annotation),
        ]

        for sequence, handler in shortcut_bindings:
            shortcut = QShortcut(QKeySequence(sequence), self)
            shortcut.setContext(Qt.WindowShortcut)
            shortcut.activated.connect(handler)
            self.shortcuts.append(shortcut)

    def choose_audio_folder(self) -> None:
        directory = QFileDialog.getExistingDirectory(self, "选择音频文件夹", str(Path.cwd()))
        if not directory:
            return
        self.load_audio_folder(Path(directory))

    def choose_excel_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 Excel 文件",
            str(Path.cwd()),
            "Excel Files (*.xlsx *.xlsm *.xltx *.xltm)",
        )
        if not file_path:
            return
        self.load_excel_mapping(Path(file_path))

    def load_excel_mapping(self, excel_path: Path) -> None:
        try:
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            QMessageBox.warning(self, "Excel 读取失败", f"无法读取 Excel：{exc}")
            return

        mapping: dict[str, dict[str, str]] = {}
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue
            if not row:
                continue

            audio_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            source_text = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            translation_text = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            if not audio_name:
                continue
            mapping[normalize_audio_key(audio_name)] = {
                "source": source_text,
                "translation": translation_text,
            }

        self.transcript_map = mapping
        self.update_transcript_display()
        QMessageBox.information(self, "导入成功", f"已读取 {len(mapping)} 条文本映射。")

    def load_audio_folder(self, folder: Path) -> None:
        files = sorted(
            path for path in folder.iterdir() if path.is_file() and path.suffix.lower() in SUPPORTED_AUDIO_EXTENSIONS
        )
        if not files:
            QMessageBox.warning(self, "未找到音频", "当前文件夹中没有支持的音频文件。")
            return

        self.audio_dir = folder
        self.audio_files = files
        self.refresh_annotated_cache()
        self.audio_index = self.first_unannotated_index()
        if self.audio_index < 0:
            self.audio_index = 0
        self.load_current_audio(auto_play=False)

    def load_current_audio(self, auto_play: bool = True) -> None:
        if not self.audio_files or self.audio_index < 0:
            return

        audio_path = self.audio_files[self.audio_index]
        self.update_audio_header(audio_path)
        self.player.setSource(QUrl.fromLocalFile(str(audio_path)))
        self.waveform_bar.set_progress(0, 0)
        self.start_waveform_decode(audio_path)
        self.time_label.setText("00:00 / 00:00")
        self.update_transcript_display()
        self.load_existing_annotation()
        self.update_summary()

        if auto_play:
            self.player.play()

    def play_previous(self) -> None:
        if not self.audio_files:
            return
        if self.audio_index > 0:
            self.audio_index -= 1
            self.load_current_audio()

    def play_next(self) -> None:
        if not self.audio_files:
            return
        if self.audio_index < len(self.audio_files) - 1:
            self.audio_index += 1
            self.load_current_audio()

    def pause_playback(self) -> None:
        self.player.pause()

    def resume_playback(self) -> None:
        if not self.audio_files:
            self.choose_audio_folder()
            return
        self.player.play()

    def toggle_playback(self) -> None:
        if not self.audio_files:
            self.choose_audio_folder()
            return
        if self.player.playbackState() == QMediaPlayer.PlayingState:
            self.player.pause()
        else:
            self.player.play()

    def _on_waveform_seek(self, ratio: float) -> None:
        duration = self.player.duration()
        if duration > 0:
            self.player.setPosition(int(duration * ratio))

    def _sync_position(self, position: int) -> None:
        self.waveform_bar.set_progress(position, self.player.duration())
        self.time_label.setText(f"{format_ms(position)} / {format_ms(self.player.duration())}")

    def _sync_duration(self, duration: int) -> None:
        self.waveform_bar.set_progress(self.player.position(), duration)
        self.time_label.setText(f"{format_ms(self.player.position())} / {format_ms(duration)}")

    def _on_media_status_changed(self, status: QMediaPlayer.MediaStatus) -> None:
        if status == QMediaPlayer.EndOfMedia:
            self.player.pause()
            self.player.setPosition(0)
            self.waveform_bar.set_progress(0, self.player.duration())
            self.time_label.setText(f"00:00 / {format_ms(self.player.duration())}")

    def current_audio_path(self) -> Path | None:
        if not self.audio_files or self.audio_index < 0:
            return None
        return self.audio_files[self.audio_index]

    def current_json_path(self) -> Path | None:
        audio_path = self.current_audio_path()
        if not audio_path:
            return None
        return self.json_dir / f"{audio_path.stem}.json"

    def refresh_annotated_cache(self) -> None:
        self.annotated_audio_ids = {path.stem for path in self.json_dir.glob("*.json")}

    def is_audio_annotated(self, audio_path: Path | None) -> bool:
        if audio_path is None:
            return False
        return audio_path.stem in self.annotated_audio_ids

    def first_unannotated_index(self) -> int:
        for index, audio_path in enumerate(self.audio_files):
            if not self.is_audio_annotated(audio_path):
                return index
        return -1

    def update_audio_header(self, audio_path: Path | None) -> None:
        if audio_path is None:
            self.audio_name_label.setText("当前音频：未加载")
            self.annotation_status_label.setText("未标注")
            self.annotation_status_label.setProperty("annotated", False)
        else:
            annotated = self.is_audio_annotated(audio_path)
            self.audio_name_label.setText(f"当前音频：{audio_path.name}")
            self.annotation_status_label.setText("✓ 已标注" if annotated else "未标注")
            self.annotation_status_label.setProperty("annotated", annotated)
        self.annotation_status_label.style().unpolish(self.annotation_status_label)
        self.annotation_status_label.style().polish(self.annotation_status_label)
        self.annotation_status_label.update()

    def update_transcript_display(self) -> None:
        audio_path = self.current_audio_path()
        if not audio_path:
            self.source_text.clear()
            self.translation_text.clear()
            return

        record = self.transcript_map.get(normalize_audio_key(audio_path.name), {})
        self.source_text.setPlainText(record.get("source", ""))
        self.translation_text.setPlainText(record.get("translation", ""))

    def _update_annotator_name(self, text: str) -> None:
        self.annotator_name = text.strip()
        self.update_summary()

    def collect_annotation(self) -> dict:
        audio_path = self.current_audio_path()
        payload: dict[str, object] = {"audio_id": audio_path.stem if audio_path else ""}

        for key, group in self.single_groups.items():
            checked = group.checkedButton()
            payload[key] = checked.property("value") if checked else ""

        for key, checkboxes in self.multi_groups.items():
            payload[key] = [box.property("value") for box in checkboxes if box.isChecked()]

        payload[REMARK_KEY] = self.remark_edit.toPlainText().strip()
        payload["annotator"] = self.annotator_name
        return payload

    def update_summary(self) -> None:
        payload = self.collect_annotation()
        self.preview_cache = json.dumps(payload, ensure_ascii=False, indent=2)

    def apply_default_selections(self) -> None:
        for key, question in self.question_configs.items():
            default_value = question.get("default")
            if question["type"] == "single":
                if not default_value:
                    continue
                group = self.single_groups[key]
                for button in group.buttons():
                    if button.property("value") == default_value:
                        button.setChecked(True)
                        break
            else:
                selected_defaults = set(default_value or [])
                for checkbox in self.multi_groups[key]:
                    checkbox.setChecked(checkbox.property("value") in selected_defaults)

        self.update_summary()

    def _handle_multi_change(self, key: str, changed_box: QCheckBox) -> None:
        exclusive_value = self.multi_exclusive_values.get(key)
        if exclusive_value is None:
            self.update_summary()
            return

        checkboxes = self.multi_groups[key]
        changed_value = changed_box.property("value")

        if changed_box.isChecked() and changed_value == exclusive_value:
            for checkbox in checkboxes:
                if checkbox is changed_box:
                    continue
                blocker = QSignalBlocker(checkbox)
                checkbox.setChecked(False)
                del blocker
        elif changed_box.isChecked() and changed_value != exclusive_value:
            for checkbox in checkboxes:
                if checkbox.property("value") == exclusive_value:
                    blocker = QSignalBlocker(checkbox)
                    checkbox.setChecked(False)
                    del blocker
                    break
        elif not any(box.isChecked() for box in checkboxes):
            for checkbox in checkboxes:
                if checkbox.property("value") == exclusive_value:
                    blocker = QSignalBlocker(checkbox)
                    checkbox.setChecked(True)
                    del blocker
                    break

        self.update_summary()

    def show_preview_dialog(self) -> None:
        self.update_summary()
        dialog = QDialog(self)
        dialog.setWindowTitle("JSON 预览")
        dialog.resize(700, 620)
        dialog.setStyleSheet(self.dialog_stylesheet())

        layout = QVBoxLayout(dialog)
        preview = QPlainTextEdit()
        preview.setObjectName("jsonPreview")
        preview.setReadOnly(True)
        preview.setPlainText(self.preview_cache)
        layout.addWidget(preview)

        close_button = QPushButton("关闭")
        close_button.setProperty("variant", "primary")
        close_button.clicked.connect(dialog.accept)

        button_row = QHBoxLayout()
        button_row.addStretch()
        button_row.addWidget(close_button)
        layout.addLayout(button_row)
        dialog.exec()

    def reset_current_annotation(self) -> None:
        for group in self.single_groups.values():
            group.setExclusive(False)
            for button in group.buttons():
                button.setChecked(False)
            group.setExclusive(True)

        for checkboxes in self.multi_groups.values():
            for checkbox in checkboxes:
                checkbox.setChecked(False)

        self.remark_edit.clear()
        self.apply_default_selections()

    def load_existing_annotation(self) -> None:
        self.reset_current_annotation()
        json_path = self.current_json_path()
        if not json_path or not json_path.exists():
            return

        try:
            payload = json.loads(json_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            QMessageBox.warning(self, "读取失败", f"已有标注文件读取失败：{json_path.name}")
            return

        for key, group in self.single_groups.items():
            value = payload.get(key)
            if not value:
                continue
            for button in group.buttons():
                if button.property("value") == value:
                    button.setChecked(True)
                    break

        for key, checkboxes in self.multi_groups.items():
            selected_values = set(payload.get(key, []))
            for checkbox in checkboxes:
                checkbox.setChecked(checkbox.property("value") in selected_values)

        blocker = QSignalBlocker(self.remark_edit)
        self.remark_edit.setPlainText(payload.get(REMARK_KEY, ""))
        del blocker
        self.update_summary()

    def submit_annotation(self) -> None:
        audio_path = self.current_audio_path()
        if not audio_path:
            QMessageBox.information(self, "提示", "请先打开一个包含音频文件的文件夹。")
            return

        missing_labels: list[str] = []
        for key in self.required_question_keys:
            group = self.single_groups.get(key)
            if group is not None and group.checkedButton() is None:
                missing_labels.append(self.question_labels[key])

        if missing_labels:
            QMessageBox.warning(self, "必填项未完成", f"请先完成这些题目：{', '.join(missing_labels)}")
            return

        payload = self.collect_annotation()
        json_path = self.current_json_path()
        if not json_path:
            return

        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        self.refresh_annotated_cache()
        self.update_audio_header(self.current_audio_path())
        QMessageBox.information(self, "提交成功", f"已生成：{json_path}")

    def toggle_theme(self) -> None:
        self.current_theme = "dark" if self.theme_button.isChecked() else "light"
        self.theme_button.setText("切换日间" if self.current_theme == "dark" else "切换夜间")
        self.apply_theme()

    def create_glass_card(
        self,
        object_name: str,
        left: int = 16,
        top: int = 16,
        right: int = 16,
        bottom: int = 16,
    ) -> tuple[QFrame, QVBoxLayout]:
        frame = QFrame()
        frame.setObjectName(object_name)
        frame.setFrameShape(QFrame.StyledPanel)

        layout = QVBoxLayout(frame)
        layout.setContentsMargins(left, top, right, bottom)
        layout.setSpacing(10)

        self.attach_shadow(frame)
        return frame, layout

    def attach_shadow(self, widget: QWidget) -> None:
        effect = QGraphicsDropShadowEffect(self)
        effect.setBlurRadius(36)
        effect.setOffset(0, 16)
        widget.setGraphicsEffect(effect)
        self.shadow_targets.append(widget)

    def apply_theme(self) -> None:
        theme = THEMES[self.current_theme]
        for widget in self.shadow_targets:
            effect = widget.graphicsEffect()
            if isinstance(effect, QGraphicsDropShadowEffect):
                shadow_color = QColor(theme["shadow"])
                shadow_color.setAlpha(theme["shadow_alpha"])
                effect.setColor(shadow_color)

        self.waveform_bar.set_theme(
            background=theme["slider_groove"],
            played=theme["accent"],
            upcoming=theme["text_muted"],
            line=theme["card_bg"],
        )
        self.setStyleSheet(self.build_stylesheet(theme))

    def start_waveform_decode(self, audio_path: Path) -> None:
        self.waveform_decode_path = audio_path
        self.waveform_samples = []
        self.waveform_bar.set_waveform([(-0.12, 0.12)] * 180)

        if self.waveform_decoder is not None:
            try:
                self.waveform_decoder.stop()
            except RuntimeError:
                pass

        decoder = QAudioDecoder(self)
        self.waveform_decoder = decoder
        decoder.setSource(QUrl.fromLocalFile(str(audio_path)))
        decoder.bufferReady.connect(self._on_waveform_buffer_ready)
        decoder.finished.connect(self._on_waveform_decode_finished)
        decoder.start()

    def _on_waveform_buffer_ready(self) -> None:
        if self.waveform_decoder is None:
            return
        buffer = self.waveform_decoder.read()
        if not buffer.isValid():
            return
        self.waveform_samples.extend(self.extract_buffer_samples(buffer))

    def _on_waveform_decode_finished(self) -> None:
        if not self.waveform_samples:
            self.waveform_bar.set_waveform([(-0.12, 0.12)] * 180)
            return
        self.waveform_bar.set_waveform(self.compress_samples_to_waveform(self.waveform_samples, 220))

    def extract_buffer_samples(self, buffer: QAudioBuffer) -> list[float]:
        fmt = buffer.format()
        channels = max(1, fmt.channelCount())
        frames = buffer.frameCount()
        if frames <= 0:
            return []

        samples = self.buffer_to_float_samples(buffer, fmt)
        if not samples:
            return []

        mono_samples: list[float] = []
        for frame_index in range(frames):
            sample_start = frame_index * channels
            frame_sum = 0.0
            for channel_index in range(channels):
                frame_sum += samples[sample_start + channel_index]
            mono_samples.append(frame_sum / channels)
        return mono_samples

    def buffer_to_float_samples(self, buffer: QAudioBuffer, fmt: QAudioFormat) -> list[float]:
        raw = buffer.data()
        sample_format = fmt.sampleFormat()

        if sample_format == QAudioFormat.SampleFormat.UInt8:
            view = raw.cast("B")
            return [(value - 128) / 128.0 for value in view]
        if sample_format == QAudioFormat.SampleFormat.Int16:
            view = raw.cast("h")
            return [value / 32768.0 for value in view]
        if sample_format == QAudioFormat.SampleFormat.Int32:
            view = raw.cast("i")
            return [value / 2147483648.0 for value in view]
        if sample_format == QAudioFormat.SampleFormat.Float:
            view = raw.cast("f")
            return [float(value) for value in view]
        return []

    def compress_samples_to_waveform(
        self, samples: list[float], target_count: int
    ) -> list[tuple[float, float]]:
        if not samples:
            return [(-0.12, 0.12)] * target_count

        total = len(samples)
        points: list[tuple[float, float]] = []
        for index in range(target_count):
            start = int(index * total / target_count)
            end = max(start + 1, int((index + 1) * total / target_count))
            segment = samples[start:end]
            seg_min = min(segment)
            seg_max = max(segment)
            if abs(seg_max - seg_min) < 0.04:
                midpoint = (seg_max + seg_min) / 2
                seg_min = midpoint - 0.02
                seg_max = midpoint + 0.02
            points.append((seg_min, seg_max))
        return points

    def build_stylesheet(self, theme: dict[str, str | int]) -> str:
        return f"""
            QWidget#appRoot {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {theme["window_start"]},
                    stop: 1 {theme["window_end"]}
                );
                color: {theme["text_primary"]};
            }}
            QFrame#heroCard {{
                background: {theme["panel_bg"]};
                border: 1px solid {theme["card_border"]};
                border-radius: 16px;
            }}
            QFrame#panelCard {{
                background: {theme["panel_alt_bg"]};
                border: 1px solid {theme["card_border"]};
                border-radius: 14px;
            }}
            QFrame#infoCard {{
                background: {theme["card_bg"]};
                border: 1px solid {theme["card_border"]};
                border-radius: 12px;
            }}
            QFrame#questionCard {{
                background: {theme["card_bg"]};
                border: 1px solid {theme["card_border"]};
                border-radius: 12px;
            }}
            QLabel#appTitle {{
                font-size: 22px;
                font-weight: 800;
                color: {theme["text_primary"]};
            }}
            QLabel#appSubtitle {{
                font-size: 13px;
                color: {theme["text_muted"]};
            }}
            QLabel#audioNameLabel {{
                font-size: 17px;
                font-weight: 700;
                color: {theme["text_primary"]};
                padding-top: 2px;
            }}
            QLabel#annotationStatusLabel {{
                font-size: 12px;
                font-weight: 700;
                color: {theme["text_muted"]};
                padding-top: 5px;
            }}
            QLabel#annotationStatusLabel[annotated="true"] {{
                color: #18a058;
            }}
            QLabel#sectionTitle {{
                font-size: 14px;
                font-weight: 700;
                color: {theme["text_primary"]};
            }}
            QLabel#miniSectionTitle {{
                font-size: 10px;
                font-weight: 700;
                color: {theme["text_primary"]};
            }}
            QLabel#compactFieldTitle {{
                font-size: 10px;
                font-weight: 700;
                color: {theme["text_primary"]};
                min-width: 36px;
            }}
            QLabel#timeLabel {{
                font-size: 13px;
                font-weight: 600;
                color: {theme["text_muted"]};
            }}
            QLabel#questionTitle {{
                font-size: 13px;
                font-weight: 700;
                color: {theme["text_primary"]};
                padding-bottom: 2px;
            }}
            QPlainTextEdit#remarkEdit,
            QPlainTextEdit#jsonPreview,
            QPlainTextEdit#transcriptViewer {{
                background: {theme["input_bg"]};
                color: {theme["text_primary"]};
                border: 1px solid {theme["input_border"]};
                border-radius: 10px;
                padding: 4px 6px;
                font-size: 11px;
            }}
            QLineEdit#annotatorEdit {{
                background: {theme["input_bg"]};
                color: {theme["text_primary"]};
                border: 1px solid {theme["input_border"]};
                border-radius: 10px;
                padding: 4px 6px;
                font-size: 11px;
            }}
            QPlainTextEdit#jsonPreview {{
                font-family: Menlo, Monaco, monospace;
            }}
            QPushButton {{
                border-radius: 12px;
                border: none;
                padding: 6px 14px;
                font-size: 12px;
                font-weight: 700;
            }}
            QPushButton[variant="primary"] {{
                background: {theme["accent"]};
                color: white;
            }}
            QPushButton[variant="primary"]:hover {{
                background: {theme["accent_hover"]};
            }}
            QPushButton[variant="secondary"] {{
                background: {theme["secondary_bg"]};
                color: {theme["text_primary"]};
                border: 1px solid {theme["card_border"]};
            }}
            QPushButton[variant="secondary"]:hover {{
                background: {theme["secondary_hover"]};
            }}
            QRadioButton#optionButton,
            QCheckBox#optionButton {{
                color: {theme["text_primary"]};
                background: transparent;
                font-size: 11px;
                spacing: 7px;
                padding-top: 1px;
                padding-bottom: 1px;
            }}
            QRadioButton#optionButton::indicator {{
                width: 16px;
                height: 16px;
                border-radius: 8px;
                border: 1px solid {theme["input_border"]};
                background: {theme["input_bg"]};
            }}
            QRadioButton#optionButton::indicator:checked {{
                width: 16px;
                height: 16px;
                border-radius: 8px;
                border: 1px solid {theme["accent"]};
                background: {theme["accent"]};
            }}
            QCheckBox#optionButton::indicator {{
                width: 14px;
                height: 14px;
                border-radius: 3px;
                border: 1px solid {theme["input_border"]};
                background: {theme["input_bg"]};
            }}
            QCheckBox#optionButton::indicator:checked {{
                background: {theme["accent"]};
                border: 1px solid {theme["accent"]};
            }}
        """

    def dialog_stylesheet(self) -> str:
        theme = THEMES[self.current_theme]
        return f"""
            QDialog {{
                background: {theme["panel_bg"]};
                color: {theme["text_primary"]};
            }}
            QPlainTextEdit#jsonPreview {{
                background: {theme["input_bg"]};
                color: {theme["text_primary"]};
                border: 1px solid {theme["input_border"]};
                border-radius: 14px;
                padding: 8px 10px;
                font-size: 13px;
                font-family: Menlo, Monaco, monospace;
            }}
            QPushButton {{
                border-radius: 12px;
                border: none;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: 700;
            }}
            QPushButton[variant="primary"] {{
                background: {theme["accent"]};
                color: white;
            }}
        """


def main() -> None:
    app = QApplication(sys.argv)
    window = AnnotationWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
